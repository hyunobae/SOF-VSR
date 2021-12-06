import io
from openpyxl import Workbook
import os

gt_path = 'C:/Users/배재현/Desktop\test\davis\2'
decoded_path = 'C:/Users/배재현/Desktop/step ablation/step4'

gt_dir = os.listdir(gt_path)
decoded_dir = os.listdir(decoded_path)

gt_dir.sort()
decoded_dir.sort()

print(gt_dir)
print(decoded_dir)

wb = Workbook()
ws = wb.active
ns = []

def put(excel, gt, decoded):
    gt_line = gt.readline()
    decoded_line = decoded.readline()
    cnt = 2
    while True:
        gt_line = gt.readline()
        decoded_line = decoded.readline()

        if not gt_line: break

        if gt_line is not '\n':
            if gt_line[0] == 'd':
                break

            if cnt>9:
                excel.append([gt_line[10:12], gt_line[14:21], decoded_line[14:21],
                      str(round(float(float(gt_line[14:21]) - float(decoded_line[14:21])), 4)) ,
                        gt_line[29:35], decoded_line[29:35]])

            else:
                excel.append([gt_line[10:11], gt_line[12:20], decoded_line[12:20],
                      str(round(float(float(gt_line[12:20]) - float(decoded_line[12:20])), 4)),
                      gt_line[27:34], decoded_line[27:34]])
            cnt += 1


for i in range(len(gt_dir)):
    gt_txt = open(gt_path + '/' + gt_dir[i], 'r')
    decoded_txt = open(decoded_path + '/' + decoded_dir[i], 'r')

    if i == 0:
        ws.title = gt_dir[i]
        ws['A1'] = 'idx'
        ws['B1'] = 'GT PSNR'
        ws['C1'] = 'DC PSNR'
        ws['D1'] = 'diff PSNR'
        ws['E1'] = 'GT SSIM'
        ws['F1'] = 'DC SSIM'
        put(ws, gt_txt, decoded_txt)

    else:
        ns.append(wb.create_sheet(gt_dir[i]))
        t = i -1
        ns[t]['A1'] = 'idx'
        ns[t]['B1'] = 'GT PSNR'
        ns[t]['C1'] = 'DC PSNR'
        ns[t]['D1'] = 'diff PSNR'
        ns[t]['E1'] = 'GT SSIM'
        ns[t]['F1'] = 'DC SSIM'
        put(ns[t], gt_txt, decoded_txt)

wb.save('davis.xlsx')








