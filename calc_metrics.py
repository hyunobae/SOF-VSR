import io
from openpyxl import Workbook

msof = open('msof.txt', "r")
sof = open('sof.txt', "r")
temp = []

excel = Workbook()
ws = excel.active
ws['A1'] = '파일명'
ws['B1'] = 'mSOF PSNR'
ws['C1'] = 'SOF PSNR'
ws['D1'] = 'diff PSNR'
ws['E1'] = 'mSOF SSIM'
ws['F1'] = 'SOF SSIM'
ws['G1'] = 'diff SSIM'

while True:
    msof_line = msof.readline()
    sof_line = sof.readline()

    if not msof_line: break

    if msof_line is not '\n':
        # print(msof_line[-7:-1]) # ssim
        # print(msof_line[-27:-20]) # psnr
        # print(msof_line[:-41]) # fname
        ws.append([msof_line[:-42], msof_line[-27:-20], sof_line[-27:-20],
                   str(round(float(float(msof_line[-27:-20]) - float(sof_line[-27:-20])), 4)),
                   msof_line[-7:], sof_line[-7:], str(round(float(float(msof_line[-7:]) - float(sof_line[-7:])), 4))])

excel.save('result.xlsx')






